import io
import copy
from decimal import Decimal
from datetime import date
import openpyxl
from openpyxl.styles import numbers
from sqlalchemy.orm import Session

from ..models import Company, MonthlyBilling
from .tariff import get_billing_calc
from .holiday_calc import third_business_day, month_abbr

TEMPLATES_DIR = "/app/guias-excel"

NUM_WORDS = {
    1: "UN", 2: "DOS", 3: "TRES", 4: "CUATRO", 5: "CINCO",
    6: "SEIS", 7: "SIETE", 8: "OCHO", 9: "NUEVE", 10: "DIEZ",
}

SPACE_PLURAL = {"A": "TIPO A", "B": "TIPO B", "C": "TIPO C", "D": "TIPO D"}


def _spaces_description(spaces) -> str:
    parts = []
    for space in spaces:
        qty = space.quantity
        num_word = NUM_WORDS.get(qty, str(qty))
        plural = "S" if qty > 1 else ""
        parts.append(f"{num_word} ({qty}) ESPACIO{plural} DE TRABAJO ({SPACE_PLURAL[space.space_type]})")
    return " Y ".join(parts)


def _clean_ceco(ceco: str) -> str:
    """Strip year suffix like ' (2025)' from CECO codes."""
    if "(" in ceco:
        return ceco[:ceco.index("(")].strip()
    return ceco.strip()


def _observacion_recaudo(company: Company, mes_abbr: str) -> str:
    contract = company.active_contract
    spaces = contract.active_spaces if contract else []
    desc = _spaces_description(spaces)
    clean_ceco = _clean_ceco(contract.ceco) if contract else ""
    return (
        f"Pago DEL SERVICIO PROG. COWORKING LABS: {desc} "
        f"CECO:{clean_ceco} [MES {mes_abbr}]. \n"
        f"Remitir comprobante de pago a: coordinador.coworking@camaradirecta.com y "
        f"tesoreria.ccb@camaradirecta.com. Genere su pago exacto sin redondear valores."
    )


def _concepto_credito(company: Company, mes_name: str) -> str:
    contract = company.active_contract
    spaces = contract.active_spaces if contract else []
    desc = _spaces_description(spaces)
    return f"CONCESIÓN DE {desc} -COWORKING LABS- MES {mes_name}."


def _observaciones_credito() -> str:
    return (
        "CTA CORRIENTE #020-834199-94-BANCOLOMBIA. "
        "Remitir comprobante de pago a: coordinador.coworking@camaradirecta.com y "
        'tesoreria.ccb@camaradirecta.com. Asunto:"PAGO__NombreEmpresa_FRA_NroFactura__MES".'
    )


def generate_recaudo(
    companies: list[Company],
    billings: dict[int, MonthlyBilling],
    billing_month: int,
    billing_year: int,
    payment_deadline: date,
) -> bytes:
    wb = openpyxl.load_workbook(f"{TEMPLATES_DIR}/orden-recaudo.xlsx")
    ws = wb["OrdRecaudo"]

    mes = month_abbr(billing_month)
    deadline_str = payment_deadline.strftime("%m/%d/%Y")

    # Clear existing data rows (row 3 onwards) preserving formatting from row 3
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for idx, company in enumerate(companies, 1):
        billing = billings.get(company.id)
        if not billing:
            continue
        contract = company.active_contract
        clean_ceco = _clean_ceco(contract.ceco) if contract else ""
        row_num = idx + 2  # data starts at row 3

        ws.cell(row_num, 1).value = str(idx)
        ws.cell(row_num, 2).value = 11036
        ws.cell(row_num, 3).value = 2
        ws.cell(row_num, 4).value = "VENDEDOR GENERAL"
        ws.cell(row_num, 5).value = company.name
        ws.cell(row_num, 6).value = int(company.nit.replace(" ", "").replace("\xa0", "")) if company.nit.strip().lstrip("\xa0").isdigit() else company.nit.strip()
        ws.cell(row_num, 7).value = clean_ceco
        ws.cell(row_num, 8).value = company.billing_email or ""
        ws.cell(row_num, 9).value = _observacion_recaudo(company, mes)
        ws.cell(row_num, 10).value = mes
        ws.cell(row_num, 11).value = 30
        ws.cell(row_num, 12).value = float(billing.daily_rate_without_iva or 0)
        ws.cell(row_num, 13).value = float(billing.subtotal_without_iva or 0)
        ws.cell(row_num, 14).value = float(billing.total_with_iva or 0)
        ws.cell(row_num, 15).value = payment_deadline
        ws.cell(row_num, 15).number_format = "MM/DD/YYYY"
        ws.cell(row_num, 16).value = company.address or ""
        ws.cell(row_num, 17).value = "SI" if company.is_affiliated else "NO"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_credito(
    companies: list[Company],
    billings: dict[int, MonthlyBilling],
    billing_month: int,
    billing_year: int,
) -> bytes:
    from .holiday_calc import month_name as get_month_name
    wb = openpyxl.load_workbook(f"{TEMPLATES_DIR}/facturacion-credito.xlsx")
    ws = wb["Detalle Facturación CWL26"]

    mes = month_abbr(billing_month)
    mes_name = get_month_name(billing_month)

    # Clear data rows starting from row 4
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for idx, company in enumerate(companies, 1):
        billing = billings.get(company.id)
        if not billing:
            continue
        contract = company.active_contract
        clean_ceco = _clean_ceco(contract.ceco) if contract else ""
        row_num = idx + 3  # data starts at row 4

        ws.cell(row_num, 2).value = idx
        ws.cell(row_num, 3).value = 11036
        ws.cell(row_num, 4).value = clean_ceco
        ws.cell(row_num, 5).value = company.fondo_emprender_ref
        ws.cell(row_num, 6).value = company.name
        ws.cell(row_num, 7).value = company.nit.strip()
        ws.cell(row_num, 8).value = _concepto_credito(company, mes_name)
        ws.cell(row_num, 9).value = mes
        ws.cell(row_num, 10).value = 30
        ws.cell(row_num, 11).value = float(billing.daily_rate_without_iva or 0)
        ws.cell(row_num, 12).value = float(billing.subtotal_without_iva or 0)
        ws.cell(row_num, 13).value = float(billing.total_with_iva or 0)
        ws.cell(row_num, 14).value = "10 días"
        ws.cell(row_num, 15).value = _observaciones_credito()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
